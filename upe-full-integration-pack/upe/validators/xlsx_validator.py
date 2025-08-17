
from __future__ import annotations
from openpyxl import load_workbook

def validate_xlsx(path: str) -> tuple[bool,str]:
    wb = load_workbook(path)
    if not wb.sheetnames:
        return (False, "no_sheets")
    return (True, "ok")
